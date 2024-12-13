import shelve
import tkinter as tk
from tkinter import simpledialog
from tkinter.ttk import Combobox
from tkinter import ttk
import os
from tkinter import Toplevel, Label, StringVar, Button, messagebox
import openpyxl
import shutil
from pathlib import Path
from tkinter import filedialog
from tkinter.ttk import Treeview


class DatabaseApp:
    def __init__(self, master):
        self.master = master
        self.master.title("LAB2")
        self.master.configure(bg='black')
        self.db_name = None
        self.fields = ["id", "name", "surname", "class", "mark"]
        self.existing_databases = []
        self.master.geometry("1300x700")

        style = ttk.Style()
        style.configure("Treeview", background="black", foreground="purple", fieldbackground="black")
        style.configure("Treeview.Heading", background="black", foreground="purple")

        self.table_frame = tk.Frame(self.master, bg='black')
        self.table_frame.pack(side=tk.TOP, padx=10, pady=50)

        self.buttons_frame = tk.Frame(self.master, bg='black')
        self.buttons_frame.pack(side=tk.TOP, padx=10, pady=10)

        self.create_widgets()

    def create_widgets(self):
        buttons = [
            ("Создать БД", self.create_database),
            ("Открыть БД", self.open_database),
            ("Удалить БД", self.delete_database),
            ("Очистить БД", self.clear_database),
            ("Сохранить БД", self.save_database),
            ("Добавить запись", self.add_record),
            ("Удалить запись", self.delete_record),
            ("Поиск", self.search),
            ("Редактировать запись", self.edit_record),
            ("Создать backup", self.create_backup),
            ("Восстановить из backup", self.restore_backup),
            ("Конвертировать", self.convert_data),
        ]

        for i, (text, command) in enumerate(buttons):
            btn = tk.Button(self.buttons_frame, text=text, command=command, width=20, fg="purple", bg="black", font="bold")
            row, col = divmod(i, 6)
            btn.grid(row=row, column=col, padx=5, pady=5)

        self.db_label = tk.Label(self.table_frame, text="Текущая база данных: ", font=(10), fg="purple", bg="black")
        self.db_label.pack(side=tk.TOP, pady=3)

        self.table_treeview = ttk.Treeview(self.table_frame, columns=self.fields, show="headings", height=15, style="Treeview")
        for field in self.fields:
            self.table_treeview.heading(field, text=field)
            self.table_treeview.column(field, anchor=tk.CENTER, width=236)
        self.table_treeview.pack(expand=True, fill="both")

        for field in self.fields:
            self.table_treeview.heading(field, text=field)
            self.table_treeview.column(field, anchor=tk.CENTER, width=236)

        self.table_treeview.pack(expand=True, fill="both")

        self.current_db = None

    def update_table(self, db_name, db):
        self.table_treeview.delete(*self.table_treeview.get_children())

        if db:
            for i in range(len(db["id"])):
                row_data = [db[field][i] for field in self.fields]
                self.table_treeview.insert("", "end", values=row_data)

        if db_name == self.current_db:
            self.current_db = db_name
            self.update_table(db_name, db)

        self.db_label.config(text=f"Текущая база данных: {db_name}" if db_name else "Текущая база данных: ")


    def create_database(self):
        new_db_name = simpledialog.askstring("Создание БД", "Введите название для новой базы данных:")
        if new_db_name:
            if self.database_exists(new_db_name):
                messagebox.showerror("Ошибка", f"БД с именем '{new_db_name}' уже существует!")
            else:
                with shelve.open(new_db_name) as new_db:
                    for field in self.fields:
                        new_db[field] = []

                messagebox.showinfo("Успех", f"БД {new_db_name} успешно создана!")
                self.existing_databases = self.get_existing_databases()
                self.update_database_combobox()

    def database_exists(self, db_name):
        db_path = Path(db_name).with_suffix(".dat")
        return db_path.is_file() or Path(db_name).is_file()

    def update_database_combobox(self):
        combo_boxes = [widget for widget in self.master.winfo_children() if isinstance(widget, ttk.Combobox)]

        for combo_box in combo_boxes:
            combo_box.set("")
            combo_box["values"] = self.existing_databases

        if self.existing_databases:
            for combo_box in combo_boxes:
                combo_box.set(self.existing_databases[-1])

    def open_database(self):
        databases = self.get_existing_databases()
        if not databases:
            messagebox.showinfo("Информация", "Нет существующих баз данных для открытия.")
            return

        dialog = tk.Toplevel(self.master, bg='black')
        dialog.title("Выбор базы данных для открытия")

        label = tk.Label(dialog, text="Выберите базу данных для открытия:", fg="purple", bg="black")
        label.pack(pady=10, padx=100)

        combo_var = tk.StringVar()
        combo = Combobox(dialog, textvariable=combo_var, values=databases)
        combo.pack(pady=10)

        ok_button = tk.Button(dialog, text="Открыть", fg="purple", bg="black", command=lambda: self.open_selected_database(dialog, combo_var.get()))
        ok_button.pack(pady=10)

    def open_selected_database(self, dialog, selected_db):
        dialog.destroy()
        if selected_db:
            try:
                with shelve.open(selected_db) as db:
                    self.update_table(selected_db, db)
            except FileNotFoundError:
                messagebox.showerror("Ошибка", "Файл базы данных не найден!")


    def show_database_content(self, db_name, db):
        content_window = tk.Toplevel(self.master)
        content_window.title(f"Содержимое БД {db_name}")

        text_widget = tk.Text(content_window, wrap="none")
        text_widget.pack(expand=True, fill="both")

        header = "\t".join(self.fields) + "\n"
        text_widget.insert(tk.END, header)

        for i in range(len(db["id"])):
            row_data = "\t".join(str(db[field][i]) for field in self.fields) + "\n"
            text_widget.insert(tk.END, row_data)

        text_widget.config(state=tk.DISABLED)

    def delete_database(self):
        databases = self.get_existing_databases()
        if not databases:
            messagebox.showinfo("Информация", "Нет существующих баз данных для удаления.")
            return

        dialog = tk.Toplevel(self.master)
        dialog.title("Выбор базы данных для удаления")

        label = tk.Label(dialog, text="Выберите базу данных для удаления:")
        label.pack(pady=10, padx=100)

        combo_var = tk.StringVar()
        combo = Combobox(dialog, textvariable=combo_var, values=databases)
        combo.pack(pady=10)

        ok_button = tk.Button(dialog, text="Удалить",
                              command=lambda: self.delete_selected_database(dialog, combo_var.get()))
        ok_button.pack(pady=10)

    def delete_selected_database(self, dialog, selected_db):
        dialog.destroy()
        if selected_db:
            try:
                os.remove(selected_db + ".dat")
                messagebox.showinfo("Успех", f"БД {selected_db} успешно удалена!")
                self.update_table(selected_db, {})
                self.current_db = None
                self.db_label.config(text="Текущая база данных:")
            except FileNotFoundError:
                messagebox.showerror("Ошибка", "Файл базы данных не найден!")

    def get_existing_databases(self, include_backups=False):
        db_files = [f for f in os.listdir() if f.endswith('.dat') and (include_backups or '_backup' not in f)]
        return [db[:-4] for db in db_files]

    def clear_database(self):
        databases = self.get_existing_databases()
        if not databases:
            messagebox.showinfo("Информация", "Нет существующих баз данных для очистки.")
            return

        dialog = tk.Toplevel(self.master)
        dialog.title("Выбор базы данных для очистки")

        label = tk.Label(dialog, text="Выберите базу данных для очистки:")
        label.pack(pady=10, padx=100)

        combo_var = tk.StringVar()
        combo = Combobox(dialog, textvariable=combo_var, values=databases)
        combo.pack(pady=10)

        ok_button = tk.Button(dialog, text="Очистить",
                              command=lambda: self.clear_selected_database(dialog, combo_var.get()))
        ok_button.pack(pady=10)

    def clear_selected_database(self, dialog, selected_db):
        dialog.destroy()
        if selected_db:
            with shelve.open(selected_db, writeback=True) as db:
                for field in self.fields:
                    db[field] = []
            messagebox.showinfo("Успех", f"БД {selected_db} успешно очищена!")
            self.update_table(selected_db, {})

    def save_database(self):
        databases = self.get_existing_databases()
        if not databases:
            messagebox.showinfo("Информация", "Нет существующих баз данных для сохранения.")
            return

        dialog = tk.Toplevel(self.master)
        dialog.title("Выбор базы данных для сохранения")

        label = tk.Label(dialog, text="Выберите базу данных для сохранения:")
        label.pack(pady=10, padx=100)

        combo_var = tk.StringVar()
        combo = ttk.Combobox(dialog, textvariable=combo_var, values=databases)
        combo.pack(pady=10)

        ok_button = tk.Button(dialog, text="Сохранить",
                              command=lambda: self.save_selected_database(dialog, combo_var.get()))
        ok_button.pack(pady=10)

    def save_selected_database(self, dialog, selected_db):
        dialog.destroy()
        if selected_db in self.get_existing_databases():
            file_path = filedialog.asksaveasfilename(defaultextension=".dat", filetypes=[("Database files", "*.dat")])
            if file_path:
                shutil.copyfile(selected_db + ".dat", file_path)
                messagebox.showinfo("Успех", f"БД {selected_db} успешно сохранена в {file_path}!")
        else:
            messagebox.showerror("Ошибка", "Выбранная база данных не существует!")

    def add_record(self):
        databases = self.get_existing_databases()
        if not databases:
            messagebox.showinfo("Информация", "Нет существующих баз данных для добавления записи.")
            return

        dialog = tk.Toplevel(self.master)
        dialog.title("Выбор базы данных для добавления записи")

        label = tk.Label(dialog, text="Выберите базу данных для добавления записи:")
        label.pack(pady=10, padx=100)

        combo_var = tk.StringVar()
        combo = Combobox(dialog, textvariable=combo_var, values=databases)
        combo.pack(pady=10)

        ok_button = tk.Button(dialog, text="Добавить запись",
                              command=lambda: self.add_record_selected_database(dialog, combo_var.get()))
        ok_button.pack(pady=10)

    def add_record_selected_database(self, dialog, selected_db):
        dialog.destroy()
        if selected_db:
            with shelve.open(selected_db, writeback=True) as db:
                record = {}
                for field in self.fields[1:]:
                    value = simpledialog.askstring("Ввод данных", f"Введите значение для поля '{field}':")
                    if value is not None:
                        record[field] = value
                    else:
                        return

                record["id"] = len(db["id"]) + 1

                for field, value in record.items():
                    db[field].append(value)

                self.update_table(selected_db, db)

            messagebox.showinfo("Успех", f"Запись успешно добавлена в БД {selected_db}!")

    def delete_record(self):
        databases = self.get_existing_databases()
        if not databases:
            messagebox.showinfo("Информация", "Нет существующих баз данных для удаления записи.")
            return

        dialog = tk.Toplevel(self.master)
        dialog.title("Выбор базы данных и id для удаления записи")

        label_db = tk.Label(dialog, text="Выберите базу данных для удаления записи:")
        label_db.pack(pady=10, padx=100)

        combo_var_db = tk.StringVar()
        combo_db = ttk.Combobox(dialog, textvariable=combo_var_db, values=databases, state="readonly")
        combo_db.pack(pady=10)

        label_id = tk.Label(dialog, text="Выберите id для удаления записи:")
        label_id.pack(pady=10, padx=100)

        combo_var_id = tk.StringVar()
        combo_id = ttk.Combobox(dialog, textvariable=combo_var_id, state="readonly")
        combo_id.pack(pady=10)

        combo_db.bind("<<ComboboxSelected>>", lambda event: self.update_id_values(combo_var_db.get(), combo_id))
        combo_db.set("")

        ok_button = ttk.Button(dialog, text="Выбрать", command=lambda: self.delete_record_selected_database(
            dialog, combo_var_db.get(), int(combo_id.get())))
        ok_button.pack(pady=10)

    def update_id_values(self, selected_db, combo_id):
        if selected_db:
            try:
                with shelve.open(selected_db) as db:
                    ids = list(map(str, db["id"]))
                    combo_id["values"] = ids
                    combo_id.set(ids[0] if ids else "")
            except FileNotFoundError:
                messagebox.showerror("Ошибка", "Файл базы данных не найден!")

    def delete_record_selected_database(self, dialog, selected_db, selected_id):
        dialog.destroy()
        if selected_db and selected_id:
            try:
                with shelve.open(selected_db, writeback=True) as db:
                    db["id"].remove(selected_id)
                    for field in self.fields[1:]:
                        db[field].pop(selected_id - 1)

                    self.update_table(selected_db, db)

                messagebox.showinfo("Успех", f"Запись с id {selected_id} успешно удалена из БД {selected_db}!")
            except FileNotFoundError:
                messagebox.showerror("Ошибка", "Файл базы данных не найден!")

    def search(self):
        databases = self.get_existing_databases()
        if not databases:
            messagebox.showinfo("Информация", "Нет существующих баз данных для поиска записей.")
            return

        dialog = tk.Toplevel(self.master)
        dialog.title("Выбор базы данных для поиска записей")

        label_db = tk.Label(dialog, text="Выберите базу данных для поиска записей:")
        label_db.pack(pady=10, padx=100)

        combo_var_db = tk.StringVar()
        combo_db = Combobox(dialog, textvariable=combo_var_db, values=databases)
        combo_db.pack(pady=10)

        label_field = tk.Label(dialog, text="Выберите поле для поиска:")
        label_field.pack(pady=10)

        combo_var_field = tk.StringVar()
        combo_field = Combobox(dialog, textvariable=combo_var_field, values=self.fields)
        combo_field.pack(pady=10)

        ok_button = tk.Button(dialog, text="Поиск",
                              command=lambda: self.search_selected_database(dialog, combo_var_db.get(),
                                                                            combo_var_field.get()))
        ok_button.pack(pady=10)

    def search_selected_database(self, dialog, selected_db, selected_field):
        dialog.destroy()
        if selected_db and selected_field:
            value = simpledialog.askstring("Поиск записей", f"Введите значение поля {selected_field} для поиска:")
            if value is not None:
                with shelve.open(selected_db) as db:
                    if selected_field in db:
                        results = []
                        for i, val in enumerate(db[selected_field]):
                            if str(val) == value:
                                result = {}
                                for field in self.fields:
                                    result[field] = db[field][i]
                                results.append(result)
                        if results:
                            self.show_search_results(results)
                        else:
                            messagebox.showinfo("Результаты поиска", "Записей не найдено!")
                    else:
                        messagebox.showerror("Ошибка", f"Указанное поле {selected_field} не найдено в базе данных!")
            else:
                messagebox.showerror("Ошибка", f"Введите значение поля {selected_field} для поиска!")

    def show_search_results(self, results):
        result_str = "Результаты поиска:\n\n"
        for result in results:
            for field, value in result.items():
                result_str += f"{field}: {value}\n"
            result_str += "\n"
        messagebox.showinfo("Результаты поиска", result_str)

    def edit_record(self):
        databases = self.get_existing_databases()
        if not databases:
            messagebox.showinfo("Информация", "Нет существующих баз данных для редактирования записи.")
            return

        db_dialog = tk.Toplevel(self.master)
        db_dialog.title("Выбор базы данных для редактирования записи")

        db_label = tk.Label(db_dialog, text="Выберите базу данных:")
        db_label.pack(pady=10, padx=100)

        db_combo_var = tk.StringVar()
        db_combo = ttk.Combobox(db_dialog, textvariable=db_combo_var, values=databases, state="readonly")
        db_combo.pack(pady=10)

        db_ok_button = ttk.Button(db_dialog, text="Выбрать",
                                  command=lambda: self.edit_record_select_id(db_dialog, db_combo_var.get()))
        db_ok_button.pack(pady=10)

    def edit_record_select_id(self, db_dialog, selected_db):
        db_dialog.destroy()
        if selected_db:
            try:
                with shelve.open(selected_db) as db:
                    ids = list(map(str, db["id"]))

                    if not ids:
                        messagebox.showinfo("Информация", f"База данных '{selected_db}' пуста.")
                        return

                    id_dialog = tk.Toplevel(self.master)
                    id_dialog.title("Выбор id записи")

                    id_label = tk.Label(id_dialog, text="Выберите id записи:")
                    id_label.pack(pady=10, padx=100)

                    id_combo_var = tk.StringVar()
                    id_combo = ttk.Combobox(id_dialog, textvariable=id_combo_var, values=ids, state="readonly")
                    id_combo.pack(pady=10)

                    id_ok_button = ttk.Button(id_dialog, text="Выбрать",
                                              command=lambda: self.edit_record_select_field(
                                                  id_dialog, selected_db, int(id_combo_var.get())))
                    id_ok_button.pack(pady=10)
            except Exception as e:
                messagebox.showerror("Ошибка", f"Произошла ошибка при открытии базы данных: {str(e)}")

    def edit_record_select_field(self, id_dialog, selected_db, selected_id):
        id_dialog.destroy()
        if selected_db and selected_id:
            try:
                with shelve.open(selected_db) as db:
                    fields_dialog = tk.Toplevel(self.master)
                    fields_dialog.title("Выбор поля для редактирования записи")

                    field_label = tk.Label(fields_dialog, text="Выберите поле для редактирования:")
                    field_label.pack(pady=10, padx=100)

                    field_values = ["все"] + self.fields[1:]
                    field_combo_var = tk.StringVar()
                    field_combo = ttk.Combobox(fields_dialog, textvariable=field_combo_var, values=field_values,
                                               state="readonly")
                    field_combo.pack(pady=10)

                    field_ok_button = ttk.Button(fields_dialog, text="Выбрать",
                                                 command=lambda: self.edit_record_enter_value(
                                                     fields_dialog, selected_db, selected_id, field_combo_var.get()))
                    field_ok_button.pack(pady=10)
            except Exception as e:
                messagebox.showerror("Ошибка", f"Произошла ошибка при открытии базы данных: {str(e)}")

    def edit_record_enter_value(self, fields_dialog, selected_db, selected_id, selected_field):
        fields_dialog.destroy()
        if selected_db and selected_id:
            try:
                with shelve.open(selected_db, writeback=True) as db:
                    if selected_field == "все":
                        new_values = {}
                        for field in self.fields[1:]:
                            new_value = simpledialog.askstring("Редактирование записи",
                                                               f"Введите новое значение для поля '{field}' "
                                                               f"и записи с id {selected_id}:",
                                                               initialvalue=db[field][db["id"].index(selected_id)])
                            if new_value is not None:
                                new_values[field] = new_value
                        for field, new_value in new_values.items():
                            db[field][db["id"].index(selected_id)] = new_value
                    else:
                        new_value = simpledialog.askstring("Редактирование записи",
                                                           f"Введите новое значение для поля '{selected_field}' "
                                                           f"и записи с id {selected_id}:",
                                                           initialvalue=db[selected_field][db["id"].index(selected_id)])
                        if new_value is not None:
                            db[selected_field][db["id"].index(selected_id)] = new_value

                    messagebox.showinfo("Успех", f"Запись с id {selected_id} успешно отредактирована!")
                    self.update_table(selected_db, db)
            except Exception as e:
                messagebox.showerror("Ошибка", f"Произошла ошибка при открытии/записи базы данных: {str(e)}")

    def create_backup(self):
        databases = self.get_existing_databases()
        if not databases:
            messagebox.showinfo("Информация", "Нет существующих баз данных для создания бэкапа.")
            return

        dialog = Toplevel(self.master)
        dialog.title("Выбор базы данных для создания бэкапа")

        label = Label(dialog, text="Выберите базу данных для создания бэкапа:")
        label.pack(pady=10, padx=100)

        combo_var = StringVar()
        combo = Combobox(dialog, textvariable=combo_var, values=databases)
        combo.pack(pady=10)

        ok_button = Button(dialog, text="Создать бэкап", command=lambda: self.create_backup_selected_database(dialog, combo_var.get()))
        ok_button.pack(pady=10)

    def create_backup_selected_database(self, dialog, selected_db):
        backup_name = f"{selected_db}_backup"
        with shelve.open(selected_db) as original_db, shelve.open(backup_name) as backup_db:
            for field in self.fields:
                backup_db[field] = original_db[field][:]

        messagebox.showinfo("Успех", f"Бэкап для базы данных {selected_db} создан: {backup_name}")
        dialog.destroy()

    def restore_backup(self):
        databases = self.get_existing_databases(include_backups=True)
        backup_databases = [db for db in databases if db.endswith("_backup")]

        if not backup_databases:
            messagebox.showinfo("Информация", "Нет существующих бэкапов для восстановления.")
            return

        dialog = tk.Toplevel(self.master)
        dialog.title("Выбор базы данных для восстановления из бэкапа")

        label = tk.Label(dialog, text="Выберите базу данных для восстановления из бэкапа:")
        label.pack(pady=10, padx=100)

        combo_var = tk.StringVar()
        combo = Combobox(dialog, textvariable=combo_var, values=backup_databases)
        combo.pack(pady=10)

        ok_button = tk.Button(dialog, text="Восстановить из бэкапа",
                              command=lambda: self.restore_backup_selected_database(dialog, combo_var.get()))
        ok_button.pack(pady=10)

    def restore_backup_selected_database(self, dialog, selected_db):
        if not selected_db.endswith("_backup"):
            messagebox.showinfo("Ошибка", "Выберите корректный файл бэкапа.")
            return

        original_name = selected_db[:-7]
        restored_data = {}
        with shelve.open(selected_db) as backup_db, shelve.open(original_name, writeback=True) as original_db:
            for field in self.fields:
                restored_data[field] = original_db[field] = backup_db[field]

        messagebox.showinfo("Успех", f"База данных восстановлена из бэкапа: {original_name}")
        dialog.destroy()

        self.update_table(original_name, restored_data)

    def convert_data(self):
        databases = self.get_existing_databases()
        if not databases:
            messagebox.showinfo("Информация", "Нет существующих баз данных для конвертации.")
            return

        dialog = tk.Toplevel(self.master)
        dialog.title("Выбор базы данных и формата конвертации")

        label_db = tk.Label(dialog, text="Выберите базу данных:")
        label_db.pack(pady=10, padx=100)

        combo_var_db = tk.StringVar()
        combo_db = Combobox(dialog, textvariable=combo_var_db, values=databases)
        combo_db.pack(pady=10)

        label_format = tk.Label(dialog, text="Выберите формат конвертации:")
        label_format.pack(pady=10)

        combo_var_format = tk.StringVar()
        combo_format = Combobox(dialog, textvariable=combo_var_format, values=['.txt', '.xlsx'])
        combo_format.pack(pady=10)

        ok_button = tk.Button(dialog, text="Конвертировать", command=lambda: self.convert_selected_data(dialog, combo_var_db.get(), combo_var_format.get()))
        ok_button.pack(pady=10)

    def convert_selected_data(self, dialog, selected_db, selected_format):
        dialog.destroy()
        if selected_db and selected_format:
            try:
                if selected_format == '.txt':
                    self.convert_dat_to_txt(selected_db)
                elif selected_format == '.xlsx':
                    self.convert_dat_to_excel(selected_db)
            except FileNotFoundError:
                messagebox.showerror("Ошибка", "Файл базы данных не найден!")

    def convert_dat_to_txt(self, selected_db):
        with shelve.open(selected_db) as db:
            txt_filename = f"{selected_db}.txt"
            with open(txt_filename, 'w') as txt_file:
                for field in self.fields:
                    txt_file.write(f"{field}: {', '.join(map(str, db[field]))}\n")
            messagebox.showinfo("Успех", f"БД {selected_db} успешно сконвертирована в {txt_filename}!")

    def convert_dat_to_excel(self, selected_db):
        with shelve.open(selected_db) as db:
            workbook = openpyxl.Workbook()
            sheet = workbook.active

            for col_num, field in enumerate(self.fields, 1):
                sheet.cell(row=1, column=col_num, value=field)

            for row_num, data_row in enumerate(zip(*[db[field] for field in self.fields]), 2):
                for col_num, value in enumerate(data_row, 1):
                    sheet.cell(row=row_num, column=col_num, value=value)

            excel_filename = f"{selected_db}.xlsx"
            workbook.save(excel_filename)

            messagebox.showinfo("Успех", f"Данные успешно сконвертированы в файл Excel: {excel_filename}")


if __name__ == "__main__":
    root = tk.Tk()
    app = DatabaseApp(root)
    root.mainloop()